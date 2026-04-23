#!/usr/bin/env python3
"""
COBie Export — extrait les données COBie 2.4 d'un fichier IFC vers Excel.
Usage : python3 cobie_export.py <fichier.ifc> [--output <fichier.xlsx>]
"""

import argparse
import sys
from pathlib import Path
from datetime import datetime

import ifcopenshell
import ifcopenshell.util.fm as fm
import ifcopenshell.util.element as element_util
import ifcopenshell.util.placement as placement_util
import ifcopenshell.util.unit as unit_util

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    sys.exit("openpyxl requis : pip install openpyxl")


# Couleurs COBie officielles
COBIE_HEADER_FILL = PatternFill("solid", fgColor="FF92D050")  # vert
COBIE_REQUIRED_FILL = PatternFill("solid", fgColor="FFFFFF00")  # jaune
COBIE_OPTIONAL_FILL = PatternFill("solid", fgColor="FFFFC000")  # orange


def style_header(cell, required: bool = True):
    cell.font = Font(bold=True)
    cell.fill = COBIE_HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _get_pset_value(entity, pset_name: str, prop_name: str) -> str:
    psets = element_util.get_psets(entity)
    pset = psets.get(pset_name, {})
    val = pset.get(prop_name, "n/a")
    return str(val) if val is not None else "n/a"


def _ifc_date(entity) -> str:
    try:
        return entity.OwnerHistory.LastModifiedDate or entity.OwnerHistory.CreationDate or ""
    except Exception:
        return ""


def _contact_email(entity) -> str:
    try:
        oh = entity.OwnerHistory
        if oh and oh.OwningUser and oh.OwningUser.ThePerson:
            person = oh.OwningUser.ThePerson
            if hasattr(person, "Addresses") and person.Addresses:
                for addr in person.Addresses:
                    if hasattr(addr, "ElectronicMailAddresses") and addr.ElectronicMailAddresses:
                        return addr.ElectronicMailAddresses[0]
        return "unknown@unknown.com"
    except Exception:
        return "unknown@unknown.com"


# ─── Feuilles COBie ────────────────────────────────────────────────────────────

def build_contact_sheet(ws, ifc: ifcopenshell.file):
    headers = ["Email", "CreatedBy", "CreatedOn", "Category", "Company",
               "Phone", "ExternalSystem", "ExternalObject", "ExternalIdentifier",
               "Department", "OrganizationCode", "GivenName", "FamilyName",
               "Street", "PostalBox", "Town", "StateRegion", "PostalCode", "Country"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(1, col, h))

    seen = set()
    row = 2
    for actor in ifc.by_type("IfcActorRole") or []:
        pass

    for person_and_org in ifc.by_type("IfcPersonAndOrganization"):
        person = person_and_org.ThePerson
        org = person_and_org.TheOrganization

        email = "unknown@unknown.com"
        phone = "n/a"
        street = town = state = postal = country = "n/a"

        if person.Addresses:
            for addr in person.Addresses:
                t = addr.is_a()
                if t == "IfcTelecomAddress":
                    if addr.ElectronicMailAddresses:
                        email = addr.ElectronicMailAddresses[0]
                    if addr.TelephoneNumbers:
                        phone = addr.TelephoneNumbers[0]
                elif t == "IfcPostalAddress":
                    street = (addr.AddressLines[0] if addr.AddressLines else "n/a")
                    town = addr.Town or "n/a"
                    state = addr.Region or "n/a"
                    postal = addr.PostalCode or "n/a"
                    country = addr.Country or "n/a"

        if email in seen:
            continue
        seen.add(email)

        given = person.GivenName or "n/a"
        family = person.FamilyName or "n/a"
        company = org.Name if org else "n/a"

        ws.append([
            email, email, datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "n/a", company, phone,
            "IFC", "IfcPersonAndOrganization", str(person_and_org.id()),
            "n/a", "n/a", given, family,
            street, "n/a", town, state, postal, country,
        ])
        row += 1


def build_facility_sheet(ws, ifc: ifcopenshell.file):
    headers = ["Name", "CreatedBy", "CreatedOn", "Category", "ProjectName",
               "SiteName", "LinearUnits", "AreaUnits", "VolumeUnits",
               "CurrencyUnit", "AreaMeasurement", "ExternalSystem",
               "ExternalProjectObject", "ExternalProjectIdentifier",
               "ExternalSiteObject", "ExternalSiteIdentifier",
               "ExternalFacilityObject", "ExternalFacilityIdentifier",
               "Description", "ProjectDescription", "SiteDescription", "Phase"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(1, col, h))

    project = ifc.by_type("IfcProject")[0] if ifc.by_type("IfcProject") else None
    site = ifc.by_type("IfcSite")[0] if ifc.by_type("IfcSite") else None
    building = ifc.by_type("IfcBuilding")[0] if ifc.by_type("IfcBuilding") else None

    linear = area = volume = "m"
    units = unit_util.get_project_unit(ifc, "LENGTHUNIT")
    if units:
        linear = getattr(units, "Name", "m") or "m"

    ws.append([
        (building.Name if building else "n/a"),
        _contact_email(building or project),
        datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        "n/a",
        (project.Name if project else "n/a"),
        (site.Name if site else "n/a"),
        linear, "m²", "m³", "EUR",
        "n/a", "IFC",
        "IfcProject", str(project.GlobalId if project else "n/a"),
        "IfcSite", str(site.GlobalId if site else "n/a"),
        "IfcBuilding", str(building.GlobalId if building else "n/a"),
        (building.Description if building else "n/a"),
        (project.Description if project else "n/a"),
        (site.Description if site else "n/a"),
        (project.Phase if project and hasattr(project, "Phase") else "n/a"),
    ])


def build_floor_sheet(ws, ifc: ifcopenshell.file):
    headers = ["Name", "CreatedBy", "CreatedOn", "Category", "ExternalSystem",
               "ExternalObject", "ExternalIdentifier", "Description",
               "Elevation", "Height"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(1, col, h))

    for storey in ifc.by_type("IfcBuildingStorey"):
        elevation = storey.Elevation if storey.Elevation is not None else "n/a"
        height = _get_pset_value(storey, "Pset_BuildingStoreyCommon", "NetHeight")
        ws.append([
            storey.Name or "n/a",
            _contact_email(storey),
            datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "Floor", "IFC", "IfcBuildingStorey", storey.GlobalId,
            storey.Description or "n/a",
            elevation, height,
        ])


def build_space_sheet(ws, ifc: ifcopenshell.file):
    headers = ["Name", "CreatedBy", "CreatedOn", "Category", "FloorName",
               "Description", "ExternalSystem", "ExternalObject",
               "ExternalIdentifier", "RoomTag", "UsableHeight",
               "GrossArea", "NetArea"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(1, col, h))

    for space in ifc.by_type("IfcSpace"):
        floor_name = "n/a"
        for rel in getattr(space, "Decomposes", []) or []:
            if rel.is_a("IfcRelAggregates") and rel.RelatingObject.is_a("IfcBuildingStorey"):
                floor_name = rel.RelatingObject.Name or "n/a"

        gross = _get_pset_value(space, "Qto_SpaceBaseQuantities", "GrossFloorArea")
        net = _get_pset_value(space, "Qto_SpaceBaseQuantities", "NetFloorArea")
        height = _get_pset_value(space, "Qto_SpaceBaseQuantities", "Height")
        category = _get_pset_value(space, "Pset_SpaceCommon", "Reference")

        ws.append([
            space.Name or "n/a",
            _contact_email(space),
            datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            category, floor_name,
            space.Description or (space.LongName or "n/a"),
            "IFC", "IfcSpace", space.GlobalId,
            space.Name or "n/a",
            height, gross, net,
        ])


def build_type_sheet(ws, ifc: ifcopenshell.file):
    headers = ["Name", "CreatedBy", "CreatedOn", "Category",
               "Description", "AssetType", "Manufacturer",
               "ModelNumber", "WarrantyGuarantorParts", "WarrantyDurationParts",
               "WarrantyGuarantorLabor", "WarrantyDurationLabor",
               "WarrantyDurationUnit", "ExtSystem", "ExtObject", "ExtIdentifier",
               "ReplacementCost", "ExpectedLife", "DurationUnit",
               "WarrantyDescription", "NominalLength", "NominalWidth",
               "NominalHeight", "ModelReference", "Shape", "Size",
               "Color", "Finish", "Grade", "Material", "Constituents",
               "Features", "AccessibilityPerformance", "CodePerformance",
               "SustainabilityPerformance"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(1, col, h))

    for type_entity in fm.get_cobie_types(ifc):
        psets = element_util.get_psets(type_entity)
        pset_type = psets.get("Pset_ManufacturerTypeInformation", {})
        pset_warranty = psets.get("Pset_Warranty", {})
        pset_asset = psets.get("COBie_Type", {})

        ws.append([
            type_entity.Name or "n/a",
            _contact_email(type_entity),
            datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            pset_asset.get("Category", "n/a"),
            type_entity.Description or "n/a",
            pset_asset.get("AssetType", "Fixed"),
            pset_type.get("Manufacturer", "n/a"),
            pset_type.get("ModelLabel", "n/a"),
            pset_warranty.get("PointOfContact", "n/a"),
            pset_warranty.get("WarrantyPeriod", "n/a"),
            pset_warranty.get("PointOfContact", "n/a"),
            pset_warranty.get("WarrantyPeriod", "n/a"),
            "Year",
            "IFC", type_entity.is_a(), type_entity.GlobalId,
            pset_asset.get("ReplacementCost", "n/a"),
            pset_asset.get("ExpectedLife", "n/a"),
            "Year",
            pset_warranty.get("WarrantyDescription", "n/a"),
            "n/a", "n/a", "n/a",
            pset_type.get("ModelReference", "n/a"),
            "n/a", "n/a", "n/a", "n/a", "n/a",
            pset_asset.get("Material", "n/a"),
            "n/a", "n/a", "n/a", "n/a", "n/a",
        ])


def build_component_sheet(ws, ifc: ifcopenshell.file):
    headers = ["Name", "CreatedBy", "CreatedOn", "TypeName", "Space",
               "Description", "ExtSystem", "ExtObject", "ExtIdentifier",
               "SerialNumber", "InstallationDate", "WarrantyStartDate",
               "TagNumber", "BarCode", "AssetIdentifier"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(1, col, h))

    for component in fm.get_cobie_components(ifc):
        type_name = "n/a"
        type_entity = element_util.get_type(component)
        if type_entity:
            type_name = type_entity.Name or "n/a"

        space_name = "n/a"
        for rel in ifc.get_inverse(component):
            if rel.is_a("IfcRelContainedInSpatialStructure"):
                if rel.RelatingStructure.is_a("IfcSpace"):
                    space_name = rel.RelatingStructure.Name or "n/a"
                    break

        psets = element_util.get_psets(component)
        pset_asset = psets.get("COBie_Component", psets.get("Pset_ManufacturerOccurrence", {}))

        ws.append([
            component.Name or "n/a",
            _contact_email(component),
            datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            type_name, space_name,
            component.Description or "n/a",
            "IFC", component.is_a(), component.GlobalId,
            pset_asset.get("SerialNumber", "n/a"),
            pset_asset.get("InstallationDate", "n/a"),
            pset_asset.get("WarrantyStartDate", "n/a"),
            pset_asset.get("TagNumber", component.Tag if hasattr(component, "Tag") else "n/a"),
            pset_asset.get("BarCode", "n/a"),
            pset_asset.get("AssetIdentifier", "n/a"),
        ])


def build_system_sheet(ws, ifc: ifcopenshell.file):
    headers = ["Name", "CreatedBy", "CreatedOn", "Category",
               "ComponentNames", "ExtSystem", "ExtObject",
               "ExtIdentifier", "Description"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(1, col, h))

    for system in ifc.by_type("IfcSystem"):
        components = []
        for rel in ifc.get_inverse(system):
            if rel.is_a("IfcRelAssignsToGroup"):
                for obj in rel.RelatedObjects:
                    if obj.Name:
                        components.append(obj.Name)

        ws.append([
            system.Name or "n/a",
            _contact_email(system),
            datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "n/a",
            ",".join(components) if components else "n/a",
            "IFC", "IfcSystem", system.GlobalId,
            system.Description or "n/a",
        ])


def build_zone_sheet(ws, ifc: ifcopenshell.file):
    headers = ["Name", "CreatedBy", "CreatedOn", "Category",
               "SpaceNames", "ExtSystem", "ExtObject", "ExtIdentifier",
               "Description"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(1, col, h))

    for zone in ifc.by_type("IfcZone"):
        spaces = []
        for rel in ifc.get_inverse(zone):
            if rel.is_a("IfcRelAssignsToGroup"):
                for obj in rel.RelatedObjects:
                    if obj.is_a("IfcSpace") and obj.Name:
                        spaces.append(obj.Name)

        ws.append([
            zone.Name or "n/a",
            _contact_email(zone),
            datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "n/a",
            ",".join(spaces) if spaces else "n/a",
            "IFC", "IfcZone", zone.GlobalId,
            zone.Description or "n/a",
        ])


def build_attribute_sheet(ws, ifc: ifcopenshell.file):
    headers = ["Name", "CreatedBy", "CreatedOn", "Category",
               "SheetName", "RowName", "Value", "Unit",
               "ExtSystem", "ExtObject", "ExtIdentifier",
               "Description", "AllowedValues"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(1, col, h))

    target_types = (
        list(fm.get_cobie_types(ifc)) +
        list(fm.get_cobie_components(ifc)) +
        list(ifc.by_type("IfcSpace")) +
        list(ifc.by_type("IfcBuildingStorey"))
    )

    sheet_map = {
        "IfcSpace": "Space",
        "IfcBuildingStorey": "Floor",
    }
    for cobie_type in fm.cobie_type_classes:
        sheet_map[cobie_type] = "Type"
    for cobie_comp in fm.cobie_component_classes:
        sheet_map[cobie_comp] = "Component"

    for entity in target_types:
        sheet_name = sheet_map.get(entity.is_a(), "Component")
        psets = element_util.get_psets(entity, should_inherit=False)
        for pset_name, props in psets.items():
            if pset_name.startswith("COBie_"):
                continue
            for prop_name, value in props.items():
                if prop_name == "id":
                    continue
                ws.append([
                    prop_name,
                    _contact_email(entity),
                    datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
                    pset_name,
                    sheet_name,
                    entity.Name or "n/a",
                    str(value) if value is not None else "n/a",
                    "n/a",
                    "IFC", pset_name, str(entity.GlobalId),
                    "n/a", "n/a",
                ])


# ─── Main ──────────────────────────────────────────────────────────────────────

SHEETS = [
    ("Contact",   build_contact_sheet),
    ("Facility",  build_facility_sheet),
    ("Floor",     build_floor_sheet),
    ("Space",     build_space_sheet),
    ("Type",      build_type_sheet),
    ("Component", build_component_sheet),
    ("System",    build_system_sheet),
    ("Zone",      build_zone_sheet),
    ("Attribute", build_attribute_sheet),
]


def export_cobie(ifc_path: str, output_path: str) -> None:
    print(f"Chargement : {ifc_path}")
    ifc = ifcopenshell.open(ifc_path)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, builder in SHEETS:
        print(f"  → Feuille {sheet_name}...")
        ws = wb.create_sheet(sheet_name)
        builder(ws, ifc)

        # Ajustement largeur colonnes
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(output_path)
    print(f"\nCOBie exporté → {output_path}")

    # Résumé
    print("\n─── Résumé ─────────────────────────────────")
    for sheet_name, _ in SHEETS:
        ws = wb[sheet_name]
        print(f"  {sheet_name:<12} {ws.max_row - 1:>5} lignes")


def main():
    parser = argparse.ArgumentParser(description="Export COBie 2.4 depuis un fichier IFC")
    parser.add_argument("ifc", help="Chemin vers le fichier IFC")
    parser.add_argument("--output", "-o", default=None,
                        help="Fichier Excel de sortie (défaut: <nom_ifc>_cobie.xlsx)")
    args = parser.parse_args()

    ifc_path = Path(args.ifc)
    if not ifc_path.exists():
        sys.exit(f"Fichier introuvable : {ifc_path}")

    output = args.output or str(ifc_path.with_suffix("")) + "_cobie.xlsx"
    export_cobie(str(ifc_path), output)


if __name__ == "__main__":
    main()
