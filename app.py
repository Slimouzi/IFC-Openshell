#!/usr/bin/env python3
"""
Interface Streamlit — COBie Export + Analyse de clashes type Solibri.
Lancement : streamlit run app.py
"""

from __future__ import annotations

import base64
import csv
import io
import json
import logging
import os
import tempfile
import urllib.parse
from pathlib import Path

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

# ── Config page ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="IFC OpenShell Tools",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
    .block-container { padding-top: 1.5rem; padding-bottom: 2rem; }
    .stMetric { background: rgba(255,255,255,0.03); padding: 10px; border-radius: 6px; }
    div[data-testid="stSidebarNav"] { display: none; }
    .severity-pill {
        display:inline-block; padding:2px 8px; border-radius:10px;
        font-size:11px; font-weight:600; color:white;
    }
    .sev-crit { background:#ff3b30; }
    .sev-maj  { background:#ff9500; }
    .sev-min  { background:#ffcc00; color:#333; }
</style>
""", unsafe_allow_html=True)

# ── Dépendances ────────────────────────────────────────────────────────────────
@st.cache_resource
def _check_deps():
    missing = []
    for pkg, mod in [("ifcopenshell", "ifcopenshell"), ("openpyxl", "openpyxl"),
                     ("ifcclash", "ifcclash.ifcclash")]:
        try:
            __import__(mod)
        except ImportError:
            missing.append(pkg)
    return missing

_missing = _check_deps()
if _missing:
    st.error(f"Dépendances manquantes : `pip install {' '.join(_missing)}`")
    st.stop()

import ifcopenshell                     # noqa: E402
import ifcopenshell.util.fm as fm      # noqa: E402
import openpyxl                         # noqa: E402
from ifcclash.ifcclash import Clasher, ClashSettings  # noqa: E402

from cobie_export import (              # noqa: E402
    SHEETS,
    build_contact_sheet, build_facility_sheet, build_floor_sheet,
    build_space_sheet, build_type_sheet, build_component_sheet,
    build_system_sheet, build_zone_sheet, build_attribute_sheet,
)
from clash_detection import (           # noqa: E402
    PRESETS, FULL_AUDIT_SEQUENCE,
    _make_clash_set,
    check_missing_reservations, check_space_coherence,
)
from clash_analyzer import (            # noqa: E402
    ClashRecord, build_records, smart_group,
    inspect_element, discipline_matrix, class_matrix, group_summary,
    export_bcf_zip, DISCIPLINE_MAP, discipline_of,
)


# ══════════════════════════════════════════════════════════════════════════════
# Title + header
# ══════════════════════════════════════════════════════════════════════════════
col_t1, col_t2 = st.columns([3, 1])
with col_t1:
    st.title("🏗️ IFC OpenShell Tools")
    st.caption("COBie 2.4 · Détection de clashes type Solibri · Visionneuse 3D Fragments/web-ifc")
with col_t2:
    if st.button("🔄 Nouveau fichier", use_container_width=True):
        st.cache_resource.clear()
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — Fichier IFC
# ══════════════════════════════════════════════════════════════════════════════
with st.expander("📁 **1. Fichier IFC**", expanded=("ifc_path" not in st.session_state)):
    data_dir = Path(__file__).parent / "data"
    local_ifcs = sorted(data_dir.glob("*.ifc"))

    tab_path, tab_data, tab_upload = st.tabs([
        "📂 Chemin absolu (recommandé)",
        "🗂️ Dossier data/",
        "⬆️ Upload navigateur",
    ])

    with tab_path:
        raw_path = st.text_input(
            "Chemin complet vers le fichier IFC",
            placeholder=f"{data_dir}/projet.ifc",
            key="input_path",
        )
    with tab_data:
        local_choice = None
        if local_ifcs:
            options = ["— choisir —"] + [f.name for f in local_ifcs]
            sel = st.selectbox("Fichiers dans data/", options, key="input_local")
            if sel != "— choisir —":
                local_choice = data_dir / sel
        else:
            st.info(f"Déposer des .ifc dans `{data_dir}` puis relancer.")
    with tab_upload:
        uploaded = st.file_uploader("Fichier IFC", type=["ifc"], key="input_upload")

    ifc_path: Path | None = None
    display_name = ""
    tmp_file = None

    if raw_path and raw_path.strip():
        cand = Path(raw_path.strip())
        if cand.exists() and cand.suffix.lower() == ".ifc":
            ifc_path = cand
            display_name = cand.name
        else:
            st.error(f"Fichier introuvable : `{raw_path}`")
    elif local_choice is not None:
        ifc_path = local_choice
        display_name = local_choice.name
    elif uploaded is not None:
        tmp_file = tempfile.NamedTemporaryFile(suffix=".ifc", delete=False)
        tmp_file.write(uploaded.read())
        tmp_file.flush()
        ifc_path = Path(tmp_file.name)
        display_name = uploaded.name

    if ifc_path:
        size_mb = ifc_path.stat().st_size / 1024 / 1024
        st.success(f"**{display_name}** — {size_mb:.1f} Mo")
        st.session_state["ifc_path"] = str(ifc_path)
        st.session_state["display_name"] = display_name

if "ifc_path" not in st.session_state:
    st.info("👆 Chargez ou sélectionnez un fichier IFC pour commencer.")
    st.stop()

ifc_path = Path(st.session_state["ifc_path"])
display_name = st.session_state.get("display_name", ifc_path.name)

@st.cache_resource(show_spinner="Chargement du fichier IFC…")
def load_ifc(path: str):
    return ifcopenshell.open(path)

ifc = load_ifc(str(ifc_path))

# ── Métriques d'accueil ────────────────────────────────────────────────────────
project   = ifc.by_type("IfcProject")
storeys   = ifc.by_type("IfcBuildingStorey")
spaces    = ifc.by_type("IfcSpace")
types_c   = list(fm.get_cobie_types(ifc))
comps_c   = list(fm.get_cobie_components(ifc))

c1, c2, c3, c4, c5, c6 = st.columns(6)
c1.metric("Schéma", ifc.schema)
c2.metric("Niveaux", len(storeys))
c3.metric("Espaces", len(spaces))
c4.metric("Types COBie", len(types_c))
c5.metric("Composants", len(comps_c))
c6.metric("Projet", (project[0].Name or "—")[:18] if project else "—")

st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# ONGLETS PRINCIPAUX
# ══════════════════════════════════════════════════════════════════════════════
main_tab_cobie, main_tab_clash = st.tabs([
    "📋 COBie Export",
    "⚠️ Clash Detection",
])


# ══════════════════════════════════════════════════════════════════════════════
# COBIE EXPORT
# ══════════════════════════════════════════════════════════════════════════════
with main_tab_cobie:
    st.subheader("Export COBie 2.4")

    col_fmt, col_sheets = st.columns([1, 2])
    with col_fmt:
        fmt = st.radio("Format", ["Excel (.xlsx)", "CSV (une feuille)"], key="cobie_fmt")
    with col_sheets:
        sheet_names = [n for n, _ in SHEETS]
        selected_sheets = st.multiselect(
            "Feuilles à inclure", sheet_names, default=sheet_names, key="cobie_sheets"
        )

    csv_sheet = None
    if fmt == "CSV (une feuille)" and selected_sheets:
        csv_sheet = st.selectbox("Feuille CSV", selected_sheets, key="cobie_csv_sheet")

    if st.button("⚙️ Générer", type="primary", use_container_width=True, key="btn_cobie_gen"):
        if not selected_sheets:
            st.warning("Sélectionnez au moins une feuille.")
        else:
            builder_map = {
                "Contact": build_contact_sheet, "Facility": build_facility_sheet,
                "Floor": build_floor_sheet, "Space": build_space_sheet,
                "Type": build_type_sheet, "Component": build_component_sheet,
                "System": build_system_sheet, "Zone": build_zone_sheet,
                "Attribute": build_attribute_sheet,
            }
            with st.spinner("Export en cours…"):
                wb = openpyxl.Workbook()
                wb.remove(wb.active)
                rc = {}
                for sn in selected_sheets:
                    b = builder_map.get(sn)
                    if not b: continue
                    ws = wb.create_sheet(sn)
                    b(ws, ifc)
                    for col in ws.columns:
                        w = max((len(str(c.value or "")) for c in col), default=10)
                        ws.column_dimensions[col[0].column_letter].width = min(w+4, 50)
                    rc[sn] = ws.max_row - 1

            st.success("Export terminé")
            cs = st.columns(min(len(rc), 5))
            for i,(sn,cnt) in enumerate(rc.items()):
                cs[i%5].metric(sn, f"{cnt} lignes")

            stem = display_name.replace(".ifc","")
            xlsx_buf = io.BytesIO(); wb.save(xlsx_buf); xlsx_buf.seek(0)

            if fmt == "Excel (.xlsx)":
                st.download_button("⬇️ Télécharger Excel COBie", data=xlsx_buf,
                    file_name=f"{stem}_cobie.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            else:
                ws_csv = wb[csv_sheet]
                buf = io.StringIO()
                w = csv.writer(buf, delimiter=";")
                for r in ws_csv.iter_rows(values_only=True):
                    w.writerow([v if v is not None else "" for v in r])
                st.download_button(f"⬇️ Télécharger {csv_sheet}.csv",
                    data=buf.getvalue().encode("utf-8-sig"),
                    file_name=f"{stem}_cobie_{csv_sheet}.csv",
                    mime="text/csv", use_container_width=True)
                rows = list(ws_csv.iter_rows(values_only=True))
                if len(rows) > 1:
                    st.dataframe(pd.DataFrame(rows[1:26], columns=rows[0]), use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# CLASH DETECTION — style Solibri Model Checker
# ══════════════════════════════════════════════════════════════════════════════
with main_tab_clash:

    # ── Sous-onglets ──────────────────────────────────────────────────────────
    sub_run, sub_dash, sub_issues, sub_detail, sub_viewer, sub_checks, sub_export = st.tabs([
        "🎯 Lancement",
        "📊 Dashboard",
        "📝 Issues",
        "🔍 Inspecteur",
        "🧊 Viewer 3D",
        "🩺 Audits IFC",
        "📤 Export",
    ])

    # ────────────────────────────────────────────────────────────────────────
    # 🎯 LANCEMENT
    # ────────────────────────────────────────────────────────────────────────
    with sub_run:
        st.subheader("Configuration de l'analyse")

        col1, col2 = st.columns([1, 2])
        with col1:
            run_mode = st.radio(
                "Mode",
                ["Audit complet (8 presets)", "Preset unique", "Personnalisé"],
                key="run_mode",
            )

        PRESET_LABELS = {
            "mep_vs_structure_hard":  "🔴 MEP vs Structure — Hard clash",
            "mep_vs_structure_soft":  "🟠 MEP vs Structure — Soft clash (50mm)",
            "mep_hvac_vs_pipe":       "🟡 MEP — Gaines CVC vs Tuyauteries",
            "mep_hvac_vs_elec":       "🟡 MEP — Gaines vs Câbles",
            "mep_pipe_vs_elec":       "🟡 MEP — Tuyaux vs Câbles",
            "mep_hvac_clearance":     "🟠 MEP — Dégagement inter-gaines (100mm)",
            "arch_vs_structure":      "🟡 Architecture vs Structure",
            "duplicate_elements":     "🟣 Doublons / superpositions",
        }

        selected_preset = None
        custom = {"mode":"collision","group_a":"","group_b":"","clearance":0.05,"tolerance":0.001}

        with col2:
            if run_mode == "Preset unique":
                selected_preset = st.selectbox(
                    "Preset", list(PRESET_LABELS.keys()),
                    format_func=lambda k: PRESET_LABELS[k],
                )
                p = PRESETS[selected_preset]
                info = f"Mode : `{p['mode']}`"
                if p.get("clearance"): info += f" · clearance={p['clearance']*1000:.0f}mm"
                if p.get("tolerance") is not None and p["mode"]=="intersection":
                    info += f" · tolerance={p['tolerance']*1000:.1f}mm"
                st.caption(info)

            elif run_mode == "Personnalisé":
                custom["mode"] = st.selectbox("Mode", ["collision","clearance","intersection"])
                custom["group_a"] = st.text_input("Groupe A (classes séparées par virgule)",
                    placeholder="IfcWall,IfcSlab,IfcColumn")
                custom["group_b"] = st.text_input("Groupe B",
                    placeholder="IfcDuctSegment,IfcPipeSegment")
                if custom["mode"] == "clearance":
                    custom["clearance"] = st.slider("Dégagement (mm)", 10, 500, 50) / 1000
                elif custom["mode"] == "intersection":
                    custom["tolerance"] = st.slider("Tolérance (mm)", 0, 50, 1) / 1000

        st.divider()

        # Options avancées
        col_opt1, col_opt2 = st.columns(2)
        with col_opt1:
            do_smart_group = st.checkbox("🔗 Smart grouping (cluster par proximité)", value=True)
            group_dist = st.slider("Rayon de cluster (m)", 0.1, 5.0, 1.0, 0.1, disabled=not do_smart_group)
        with col_opt2:
            do_reservations = st.checkbox("🩺 Vérifier les réservations manquantes", value=False)
            do_spatial = st.checkbox("🏠 Vérifier la cohérence spatiale", value=False)

        if st.button("🚀 Lancer l'analyse", type="primary", use_container_width=True, key="btn_run_clash"):
            with st.spinner("Analyse géométrique en cours — patientez…"):
                tmp_json = tempfile.NamedTemporaryFile(suffix=".json", delete=False)
                tmp_json.close()

                if run_mode == "Audit complet (8 presets)":
                    clash_sets = [
                        _make_clash_set(label, str(ifc_path), str(ifc_path), PRESETS[pk])
                        for pk, label in FULL_AUDIT_SEQUENCE
                    ]
                elif run_mode == "Preset unique":
                    p = PRESETS[selected_preset]
                    clash_sets = [_make_clash_set(
                        PRESET_LABELS[selected_preset], str(ifc_path), str(ifc_path), p
                    )]
                else:
                    ga = [c.strip() for c in custom["group_a"].split(",") if c.strip()]
                    gb = [c.strip() for c in custom["group_b"].split(",") if c.strip()]
                    cp = {"mode":custom["mode"], "group_a":ga, "group_b":gb,
                          "clearance":custom["clearance"], "tolerance":custom["tolerance"],
                          "allow_touching":False}
                    clash_sets = [_make_clash_set("Personnalisé", str(ifc_path), str(ifc_path), cp)]

                settings = ClashSettings()
                settings.output = tmp_json.name
                settings.logger = logging.getLogger("ifcclash")
                clasher = Clasher(settings)
                clasher.clash_sets = clash_sets
                clasher.clash()
                records = build_records(clasher.clash_sets)

                if do_smart_group and records:
                    records = smart_group(records, max_distance=group_dist)

                # Audits non-géométriques
                reservations = check_missing_reservations(ifc) if do_reservations else []
                spatial_issues = check_space_coherence(ifc) if do_spatial else []

                try: os.unlink(tmp_json.name)
                except Exception: pass

            # Persistance en session
            st.session_state["records"] = records
            st.session_state["reservations"] = reservations
            st.session_state["spatial_issues"] = spatial_issues
            # Init/reset des statuts/commentaires
            if "statuses" not in st.session_state:
                st.session_state["statuses"] = {}
                st.session_state["comments"] = {}
            for r in records:
                st.session_state["statuses"].setdefault(r.id, "open")
                st.session_state["comments"].setdefault(r.id, "")

            st.success(f"✅ {len(records)} clash(s) détecté(s) — voir les onglets Dashboard / Issues / Viewer 3D")

    # Récupération des données
    records: list[ClashRecord] = st.session_state.get("records", [])
    statuses: dict = st.session_state.get("statuses", {})
    comments: dict = st.session_state.get("comments", {})

    # DataFrame principal
    def records_to_df(recs: list[ClashRecord]) -> pd.DataFrame:
        if not recs:
            return pd.DataFrame()
        rows = []
        for r in recs:
            d = r.to_dict()
            d["status"] = statuses.get(r.id, "open")
            d["comment"] = comments.get(r.id, "")
            rows.append(d)
        return pd.DataFrame(rows)

    df_all = records_to_df(records)

    # ────────────────────────────────────────────────────────────────────────
    # 📊 DASHBOARD
    # ────────────────────────────────────────────────────────────────────────
    with sub_dash:
        if df_all.empty:
            st.info("Lancez une analyse dans l'onglet **🎯 Lancement**.")
        else:
            # KPIs
            total = len(df_all)
            sev_counts = df_all["severity"].value_counts()
            status_counts = df_all["status"].value_counts()

            k1, k2, k3, k4, k5, k6 = st.columns(6)
            k1.metric("Total", total)
            k2.metric("🔴 Critique", int(sev_counts.get("Critique", 0)))
            k3.metric("🟠 Majeur",   int(sev_counts.get("Majeur", 0)))
            k4.metric("🟡 Mineur",   int(sev_counts.get("Mineur", 0)))
            k5.metric("🟢 Résolus",  int(status_counts.get("resolved", 0)))
            k6.metric("🚫 Ignorés",  int(status_counts.get("ignored", 0)))

            st.divider()

            # Répartitions
            g1, g2 = st.columns(2)
            with g1:
                st.write("**Par sévérité**")
                st.bar_chart(df_all["severity"].value_counts())
            with g2:
                st.write("**Par type de clash**")
                st.bar_chart(df_all["type"].value_counts())

            g3, g4 = st.columns(2)
            with g3:
                st.write("**Par ensemble / preset**")
                st.bar_chart(df_all["set_name"].value_counts())
            with g4:
                st.write("**Par statut**")
                st.bar_chart(df_all["status"].value_counts())

            st.divider()
            st.write("### 🔥 Matrice disciplines A × B")
            dm = discipline_matrix(records)
            if dm:
                disciplines = sorted(set([k[0] for k in dm] + [k[1] for k in dm]))
                matrix = pd.DataFrame(0, index=disciplines, columns=disciplines)
                for (a, b), cnt in dm.items():
                    matrix.loc[a, b] = cnt
                    if a != b: matrix.loc[b, a] = cnt
                st.dataframe(
                    matrix.style.background_gradient(cmap="Reds"),
                    use_container_width=True,
                )

            st.write("### 🔝 Top 15 des paires de classes IFC")
            cm = class_matrix(records, top_k=15)
            if cm:
                df_cm = pd.DataFrame([
                    {"Classe A": k[0], "Classe B": k[1], "Clashs": v}
                    for k, v in cm.items()
                ])
                st.dataframe(df_cm, use_container_width=True, hide_index=True)

            # Groupes
            groups = group_summary(records)
            if groups:
                st.divider()
                st.write(f"### 🔗 Clusters spatiaux ({len(groups)} groupe(s))")
                df_groups = pd.DataFrame([
                    {
                        "Groupe": f"#{g['group_id']}",
                        "Clashs": g["count"],
                        "🔴 Crit": g["critical"],
                        "🟠 Maj": g["major"],
                        "🟡 Min": g["minor"],
                        "Position X": round(g["center"][0], 2),
                        "Position Y": round(g["center"][1], 2),
                        "Position Z": round(g["center"][2], 2),
                    }
                    for g in sorted(groups, key=lambda x: -x["count"])
                ])
                st.dataframe(df_groups, use_container_width=True, hide_index=True)

    # ────────────────────────────────────────────────────────────────────────
    # 📝 ISSUES — liste interactive avec filtres avancés + statut
    # ────────────────────────────────────────────────────────────────────────
    with sub_issues:
        if df_all.empty:
            st.info("Aucune analyse exécutée.")
        else:
            # Filtres avancés
            with st.expander("🔎 **Filtres avancés**", expanded=True):
                fc1, fc2, fc3 = st.columns(3)
                with fc1:
                    f_sev = st.multiselect("Sévérité",
                        ["Critique","Majeur","Mineur"],
                        default=["Critique","Majeur","Mineur"], key="f_sev")
                    f_status = st.multiselect("Statut",
                        ["open","wip","resolved","ignored"],
                        default=["open","wip"], key="f_status")
                with fc2:
                    f_type = st.multiselect("Type de clash",
                        sorted(df_all["type"].unique()),
                        default=list(df_all["type"].unique()), key="f_type")
                    f_set = st.multiselect("Ensemble / preset",
                        sorted(df_all["set_name"].unique()),
                        default=list(df_all["set_name"].unique()), key="f_set")
                with fc3:
                    all_disciplines = sorted(set(df_all["a_discipline"]) | set(df_all["b_discipline"]))
                    f_discipline = st.multiselect("Discipline (A ou B)",
                        all_disciplines, default=[], placeholder="Toutes", key="f_disc")
                    all_classes = sorted(set(df_all["a_class"]) | set(df_all["b_class"]))
                    f_class = st.multiselect("Classe IFC (A ou B)",
                        all_classes, default=[], placeholder="Toutes", key="f_cls")

                fc4, fc5, fc6 = st.columns([2, 1, 1])
                with fc4:
                    search = st.text_input("🔎 Recherche (nom / GUID / commentaire)", key="f_search")
                with fc5:
                    groups_available = sorted([int(g) for g in df_all["group_id"].dropna().unique()])
                    if groups_available:
                        f_group = st.selectbox("Groupe", ["Tous"] + [f"#{g}" for g in groups_available],
                                               key="f_group")
                    else:
                        f_group = "Tous"
                with fc6:
                    sort_col = st.selectbox("Trier par",
                        ["severity","type","distance","set_name","a_class","b_class"],
                        key="f_sort")

            # Application des filtres
            mask = (
                df_all["severity"].isin(f_sev) &
                df_all["status"].isin(f_status) &
                df_all["type"].isin(f_type) &
                df_all["set_name"].isin(f_set)
            )
            if f_discipline:
                mask &= df_all["a_discipline"].isin(f_discipline) | df_all["b_discipline"].isin(f_discipline)
            if f_class:
                mask &= df_all["a_class"].isin(f_class) | df_all["b_class"].isin(f_class)
            if f_group != "Tous":
                gid = int(f_group.replace("#",""))
                mask &= df_all["group_id"] == gid
            if search:
                s = search.lower()
                mask &= (
                    df_all["a_name"].str.lower().str.contains(s, na=False) |
                    df_all["b_name"].str.lower().str.contains(s, na=False) |
                    df_all["a_guid"].str.lower().str.contains(s, na=False) |
                    df_all["b_guid"].str.lower().str.contains(s, na=False) |
                    df_all["comment"].str.lower().str.contains(s, na=False)
                )

            df = df_all[mask].copy()
            sev_order = {"Critique":0,"Majeur":1,"Mineur":2}
            if sort_col == "severity":
                df["_o"] = df["severity"].map(sev_order)
                df = df.sort_values("_o").drop(columns="_o")
            else:
                df = df.sort_values(sort_col)

            st.caption(f"**{len(df)} / {len(df_all)}** clash(s) affiché(s)")

            # Tableau éditable (statut + commentaires)
            display_cols = ["severity","status","type","set_name",
                            "a_class","a_name","b_class","b_name",
                            "distance","group_id","comment","id"]

            edited = st.data_editor(
                df[display_cols],
                use_container_width=True,
                height=520,
                column_config={
                    "severity": st.column_config.TextColumn("Sév.", width="small", disabled=True),
                    "status": st.column_config.SelectboxColumn(
                        "Statut", options=["open","wip","resolved","ignored"], width="small",
                    ),
                    "type": st.column_config.TextColumn("Type", width="small", disabled=True),
                    "set_name": st.column_config.TextColumn("Set", width="medium", disabled=True),
                    "a_class": st.column_config.TextColumn("Classe A", disabled=True),
                    "a_name":  st.column_config.TextColumn("Nom A", disabled=True),
                    "b_class": st.column_config.TextColumn("Classe B", disabled=True),
                    "b_name":  st.column_config.TextColumn("Nom B", disabled=True),
                    "distance": st.column_config.NumberColumn("Dist.", format="%.3f m", disabled=True),
                    "group_id": st.column_config.NumberColumn("Groupe", width="small", disabled=True),
                    "comment": st.column_config.TextColumn("Commentaire", width="medium"),
                    "id": st.column_config.TextColumn("ID", width="small", disabled=True),
                },
                hide_index=True,
                key="issues_editor",
            )

            # Sauvegarde des changements (statut + commentaires)
            for _, row in edited.iterrows():
                rid = row["id"]
                if statuses.get(rid) != row["status"]:
                    statuses[rid] = row["status"]
                if comments.get(rid) != row["comment"]:
                    comments[rid] = row["comment"]
            st.session_state["statuses"] = statuses
            st.session_state["comments"] = comments

            # Actions rapides en masse
            st.write("**Actions rapides**")
            b1, b2, b3, b4 = st.columns(4)
            with b1:
                if st.button("✅ Tous → Résolu", use_container_width=True):
                    for rid in df["id"]: statuses[rid] = "resolved"
                    st.rerun()
            with b2:
                if st.button("🚫 Tous → Ignoré", use_container_width=True):
                    for rid in df["id"]: statuses[rid] = "ignored"
                    st.rerun()
            with b3:
                if st.button("🔄 Tous → Ouvert", use_container_width=True):
                    for rid in df["id"]: statuses[rid] = "open"
                    st.rerun()
            with b4:
                # Sélectionner un clash pour l'inspecteur
                if not df.empty:
                    st.session_state.setdefault("selected_clash_id", df.iloc[0]["id"])
                    options = [f"{r.severity} · {r.type} · {r.a_class[:20]} vs {r.b_class[:20]}"
                               for r in [rec for rec in records if rec.id in df["id"].values]]
                    ids = [rec.id for rec in records if rec.id in df["id"].values]
                    if ids:
                        idx = ids.index(st.session_state["selected_clash_id"]) if \
                            st.session_state["selected_clash_id"] in ids else 0
                        sel = st.selectbox("Clash à inspecter", range(len(ids)),
                                           format_func=lambda i: options[i], index=idx,
                                           key="sel_inspect")
                        st.session_state["selected_clash_id"] = ids[sel]

    # ────────────────────────────────────────────────────────────────────────
    # 🔍 INSPECTEUR
    # ────────────────────────────────────────────────────────────────────────
    with sub_detail:
        sel_id = st.session_state.get("selected_clash_id")
        rec = next((r for r in records if r.id == sel_id), None)

        if not rec:
            st.info("Sélectionnez un clash dans l'onglet **📝 Issues**.")
        else:
            # Header du clash
            sev_class = "sev-crit" if rec.severity=="Critique" else "sev-maj" if rec.severity=="Majeur" else "sev-min"
            st.markdown(
                f"### <span class='severity-pill {sev_class}'>{rec.severity}</span> "
                f"{rec.clash_type} — <code>{rec.id}</code>",
                unsafe_allow_html=True,
            )
            col_meta1, col_meta2, col_meta3, col_meta4 = st.columns(4)
            col_meta1.metric("Distance", f"{rec.distance:.4f} m")
            col_meta2.metric("Ensemble", rec.set_name[:20])
            col_meta3.metric("Statut", statuses.get(rec.id, "open"))
            col_meta4.metric("Groupe", f"#{rec.group_id}" if rec.group_id else "—")

            st.caption(f"📍 Position approx. : ({rec.center()[0]:.2f}, {rec.center()[1]:.2f}, {rec.center()[2]:.2f})")

            # Commentaire
            new_comment = st.text_area(
                "💬 Commentaire", value=comments.get(rec.id, ""), key=f"com_{rec.id}"
            )
            if new_comment != comments.get(rec.id, ""):
                comments[rec.id] = new_comment
                st.session_state["comments"] = comments

            st.divider()

            # Deux panneaux côte à côte pour A et B
            colA, colB = st.columns(2)

            for col, side in [(colA, "A"), (colB, "B")]:
                with col:
                    st.write(f"### Élément {side}")
                    guid = rec.a_guid if side == "A" else rec.b_guid
                    klass = rec.a_class if side == "A" else rec.b_class
                    name = rec.a_name if side == "A" else rec.b_name
                    disc = rec.a_discipline if side == "A" else rec.b_discipline

                    st.markdown(f"**{klass}** · _{disc}_")
                    st.markdown(f"**Nom :** `{name}`")
                    st.markdown(f"**GUID :** `{guid}`")

                    if guid:
                        info = inspect_element(ifc, guid)
                        if "error" in info:
                            st.error(info["error"])
                        else:
                            with st.expander("📋 Informations de base"):
                                for k in ["type_name","type_class","container_class",
                                          "container_name","tag","description"]:
                                    if info.get(k):
                                        st.write(f"**{k}** : {info[k]}")
                                if info.get("materials"):
                                    st.write(f"**Matériaux** : {', '.join(info['materials'])}")
                                if info.get("location"):
                                    loc = info["location"]
                                    st.write(f"**Position** : ({loc[0]:.2f}, {loc[1]:.2f}, {loc[2]:.2f})")

                            if info.get("psets"):
                                with st.expander(f"🏷️ Property Sets ({len(info['psets'])})"):
                                    for ps_name, ps_data in info["psets"].items():
                                        st.markdown(f"**`{ps_name}`**")
                                        st.json(ps_data, expanded=False)

    # ────────────────────────────────────────────────────────────────────────
    # 🧊 VIEWER 3D — Three.js + web-ifc
    # ────────────────────────────────────────────────────────────────────────
    with sub_viewer:
        if df_all.empty:
            st.info("Lancez une analyse d'abord.")
        else:
            st.caption(
                "La visionneuse utilise **web-ifc + Three.js** (chargement côté navigateur). "
                "Chargez le même fichier IFC dans le viewer via le sélecteur de fichier, "
                "les clashs apparaissent en sphères colorées par sévérité."
            )

            # Filtres de markers
            col_v1, col_v2 = st.columns([2, 1])
            with col_v1:
                v_sev = st.multiselect("Markers à afficher — sévérité",
                    ["Critique","Majeur","Mineur"],
                    default=["Critique","Majeur"], key="v_sev")
            with col_v2:
                v_focus = None
                sel_id = st.session_state.get("selected_clash_id")
                if sel_id:
                    focus_on = st.checkbox("🎯 Focus sur le clash sélectionné", value=False, key="v_focus")
                    if focus_on: v_focus = sel_id

            # Préparation des markers (JSON compacté)
            filtered = df_all[df_all["severity"].isin(v_sev) & df_all["status"].isin(["open","wip"])]
            markers_payload = []
            for _, r in filtered.iterrows():
                markers_payload.append({
                    "id": r["id"],
                    "severity": r["severity"],
                    "type": r["type"],
                    "a_class": r["a_class"], "a_name": r["a_name"],
                    "b_class": r["b_class"], "b_name": r["b_name"],
                    "distance": float(r["distance"]),
                    "center": r["center"],
                    "set_name": r["set_name"],
                })

            # Charger viewer.html et injecter les markers
            viewer_html_path = Path(__file__).parent / "viewer.html"
            viewer_html = viewer_html_path.read_text()

            # Injection : on remplace la récupération depuis URLSearchParams
            # par une constante JS afin d'éviter les limites de taille d'URL
            markers_js = json.dumps(markers_payload, ensure_ascii=False)
            focus_js = f'"{v_focus}"' if v_focus else "null"

            inject = f"""
            <script>
              window.__CLASH_MARKERS__ = {markers_js};
              window.__FOCUS_ID__ = {focus_js};
            </script>
            """
            viewer_html = viewer_html.replace(
                "const markersData = urlParams.get('markers');",
                "const markersData = null; let markers = window.__CLASH_MARKERS__ || [];",
            ).replace(
                "try {\n  if (markersData) markers = JSON.parse(decodeURIComponent(markersData));\n} catch(e) { console.warn('Markers invalides', e); }",
                "",
            ).replace(
                "const focusId = urlParams.get('focus');",
                "const focusId = window.__FOCUS_ID__;",
            ).replace(
                "</head>",
                inject + "</head>",
            )

            st.caption(f"🎯 {len(markers_payload)} marker(s) envoyé(s) à la visionneuse")
            components.html(viewer_html, height=720, scrolling=False)

            with st.expander("ℹ️ Comment utiliser la visionneuse"):
                st.markdown("""
                1. **Charger l'IFC** : utilisez le sélecteur en haut à gauche pour charger le même fichier
                2. **Naviguer** : clic-gauche pour orbiter, molette pour zoomer, clic-droit pour déplacer
                3. **Clash** : cliquez sur une sphère pour voir les infos (panneau en haut à droite)
                4. **Vue** : bouton 🎯 Recentrer, 👁️ Markers on/off, 🧬 filaire on/off
                5. **Sévérité** : 🔴 critique · 🟠 majeur · 🟡 mineur
                """)

    # ────────────────────────────────────────────────────────────────────────
    # 🩺 AUDITS IFC (réservations + cohérence spatiale)
    # ────────────────────────────────────────────────────────────────────────
    with sub_checks:
        reservations = st.session_state.get("reservations", [])
        spatial_issues = st.session_state.get("spatial_issues", [])

        st.subheader("🩺 Réservations manquantes")
        if not reservations:
            st.caption("Activez la case 'Vérifier les réservations' lors du lancement pour voir ici.")
        else:
            st.write(f"**{len(reservations)}** problème(s) détecté(s)")
            df_r = pd.DataFrame(reservations)
            st.dataframe(df_r, use_container_width=True, hide_index=True)

        st.divider()
        st.subheader("🏠 Cohérence spatiale")
        if not spatial_issues:
            st.caption("Activez la case 'Vérifier la cohérence' lors du lancement pour voir ici.")
        else:
            st.write(f"**{len(spatial_issues)}** problème(s) détecté(s)")
            df_s = pd.DataFrame(spatial_issues)
            st.dataframe(df_s, use_container_width=True, hide_index=True)

    # ────────────────────────────────────────────────────────────────────────
    # 📤 EXPORT
    # ────────────────────────────────────────────────────────────────────────
    with sub_export:
        if df_all.empty:
            st.info("Lancez une analyse d'abord.")
        else:
            st.subheader("Exports")
            stem = display_name.replace(".ifc","")

            # Re-applique les statuts/commentaires actuels sur les records
            for r in records:
                r.status = statuses.get(r.id, "open")
                r.comment = comments.get(r.id, "")

            # Filtres d'export
            export_open_only = st.checkbox("Exporter uniquement les clashs non résolus", value=False)
            recs_to_export = [r for r in records if not export_open_only
                              or r.status in ("open","wip")]

            st.caption(f"{len(recs_to_export)} / {len(records)} clash(s) à exporter")

            e1, e2, e3, e4 = st.columns(4)

            with e1:
                csv_df = pd.DataFrame([r.to_dict() for r in recs_to_export])
                csv_buf = csv_df.to_csv(index=False, sep=";").encode("utf-8-sig")
                st.download_button("📄 CSV", data=csv_buf,
                    file_name=f"{stem}_clashes.csv", mime="text/csv",
                    use_container_width=True)

            with e2:
                json_buf = json.dumps(
                    [r.to_dict() for r in recs_to_export],
                    indent=2, ensure_ascii=False, default=str,
                ).encode("utf-8")
                st.download_button("📋 JSON", data=json_buf,
                    file_name=f"{stem}_clashes.json", mime="application/json",
                    use_container_width=True)

            with e3:
                # BCF
                if st.button("🏗️ Générer BCF", use_container_width=True):
                    bcf_path = tempfile.NamedTemporaryFile(suffix=".bcfzip", delete=False)
                    bcf_path.close()
                    export_bcf_zip(recs_to_export, bcf_path.name,
                        project_name=display_name)
                    with open(bcf_path.name, "rb") as f:
                        st.session_state["bcf_data"] = f.read()
                    os.unlink(bcf_path.name)
                if "bcf_data" in st.session_state:
                    st.download_button("⬇️ Télécharger BCF",
                        data=st.session_state["bcf_data"],
                        file_name=f"{stem}_clashes.bcfzip",
                        mime="application/zip",
                        use_container_width=True)

            with e4:
                # Excel multi-feuilles
                if st.button("📊 Générer Excel", use_container_width=True):
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Clashes"
                    headers = list(recs_to_export[0].to_dict().keys()) if recs_to_export else []
                    ws.append(headers)
                    for r in recs_to_export:
                        d = r.to_dict()
                        ws.append([str(d[h]) if not isinstance(d[h], (int,float,str,bool,type(None)))
                                   else d[h] for h in headers])
                    buf = io.BytesIO()
                    wb.save(buf); buf.seek(0)
                    st.session_state["xlsx_data"] = buf.getvalue()
                if "xlsx_data" in st.session_state:
                    st.download_button("⬇️ Télécharger Excel",
                        data=st.session_state["xlsx_data"],
                        file_name=f"{stem}_clashes.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)

            st.divider()
            st.subheader("💾 Sauvegarde de l'état")

            state_data = {
                "display_name": display_name,
                "records": [r.to_dict() for r in records],
                "statuses": statuses,
                "comments": comments,
                "reservations": st.session_state.get("reservations", []),
                "spatial_issues": st.session_state.get("spatial_issues", []),
            }
            state_json = json.dumps(state_data, indent=2, ensure_ascii=False, default=str).encode("utf-8")
            st.download_button(
                "💾 Sauvegarder la session (JSON)",
                data=state_json,
                file_name=f"{stem}_clash_session.json",
                mime="application/json",
                use_container_width=True,
                help="Contient les clashs + statuts + commentaires. Rechargeable dans une future version.",
            )


# ── Cleanup ────────────────────────────────────────────────────────────────────
if tmp_file:
    try: os.unlink(tmp_file.name)
    except Exception: pass
